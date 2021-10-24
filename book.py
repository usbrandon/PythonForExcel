from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('NewSheet')
# Set this sheet to be ahead of all the others by index = 0
ws2 = wb.create_sheet('Another', 0)

ws.title = 'MySheet'

print(wb.sheetnames)

wb2 = load_workbook('regions.xlsx')

new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

cell = active_sheet['A1']
print(cell.value)