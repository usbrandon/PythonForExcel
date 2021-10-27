"""
This example will go from Pandas dataframe directly to openpyxl

"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import itertools


wb = load_workbook('regions.xlsx')
ws = wb.active
df = pd.read_excel('../Multiple_Sheets/all_shifts.xlsx')

df1 = df[['Sales Rep', 'Cost per', 'Units Sold']]
df1['Total'] = df1['Cost per'] * df1['Units Sold']
print(df1)

rows = dataframe_to_rows(df1, index=False)
print(rows) # Nope... Is an object with multiple columns and rows.

# The little lesson here is that once an iterator like the one in 'rows' is used
# You cannot use it again.  You can create other iterators using itertools to
# pass over the same dataset twice.  rows below gets used up when creating iterators
it1, it2, itr3 = itertools.tee(rows, 3)
# Data printed in lists.
for row in it1:
    print(row)

for row in it2:
    for col in row:
        print(col)

for r_idx, row in enumerate(itr3,1):
    for c_idx, col in enumerate(row,6):
        print(col)
        ws.cell(row=r_idx, column=c_idx, value=col)

wb.save('combined.xlsx')